# -*- coding: utf-8 -*-
"""
RDM Experiment Execution Wrapper for DVC Pipeline

This script executes only the RDM experiment by temporarily modifying the
Interface_RDM.xlsx configuration to run only the RDM analysis.

Usage:
    python scripts/run_rdm_experiment.py

Manual execution:
    This script can be run independently for debugging or testing RDM experiments.

DVC integration:
    Called automatically by DVC when executing the 'rdm_experiment' stage.

Author: AFR_RDM Team
"""

import os
import sys
import json
import time
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Add src to path to import RUN_RDM
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

def backup_interface(interface_path):
    """Create a backup of Interface_RDM.xlsx"""
    backup_path = interface_path.with_suffix('.xlsx.backup')
    shutil.copy2(interface_path, backup_path)
    return backup_path

def restore_interface(interface_path, backup_path):
    """Restore Interface_RDM.xlsx from backup"""
    if backup_path.exists():
        shutil.copy2(backup_path, interface_path)
        backup_path.unlink()

def configure_for_rdm_only(interface_path):
    """
    Modify Interface_RDM.xlsx to run only RDM experiment.
    Sets Run_Base_Future=No and Run_RDM=Yes
    Also caches formula values in ALL sheets to avoid NaN issues.

    IMPORTANT: openpyxl cannot evaluate formulas. When it re-saves the workbook,
    every formula cell is written WITHOUT its calculated result, so any tool that
    later reads the file (pandas) sees NaN for those cells. We therefore snapshot
    the calculated values of every formula cell (data_only=True) and write them
    back as literals before saving.

    This must cover ALL sheets, not just Uncertainty_Table. In particular the
    'Params_Sets_Vari' sheet stores the number of indexing sets per parameter in a
    '=COUNTA(...)' formula row; if that is lost to NaN, 0_experiment_manager.py
    reads df_Params_Sets_Vari.loc['Number', param] as NaN, every
    `number_sets_by_param == 1/2/3` test in PART 3 fails, no parameter is ever
    perturbed, and all futures come out identical (only Future.ID differs) ->
    PRIM gets zero-variance metrics and writes an empty CSV.
    """
    print("📝 Configuring Interface_RDM.xlsx for RDM experiment only...")

    # Load with data_only=True to get calculated values from formulas (ALL sheets)
    wb_data = load_workbook(interface_path, data_only=True)
    cached_values = {}
    for sheet_name in wb_data.sheetnames:
        ws_data = wb_data[sheet_name]
        for row_idx in range(1, ws_data.max_row + 1):
            for col_idx in range(1, ws_data.max_column + 1):
                cell_value = ws_data.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    cached_values[(sheet_name, row_idx, col_idx)] = cell_value
    wb_data.close()

    # Load normally to modify
    wb = load_workbook(interface_path)

    # Modify Setup sheet
    ws_setup = wb['Setup']
    headers = {}
    for col_idx, cell in enumerate(ws_setup[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

    if 'Run_Base_Future' in headers:
        ws_setup.cell(row=2, column=headers['Run_Base_Future'], value='No')
    if 'Run_RDM' in headers:
        ws_setup.cell(row=2, column=headers['Run_RDM'], value='Yes')

    # Write cached values back to every formula cell (all sheets) to preserve
    # formula results. The startswith('=') guard means only formula cells are
    # overwritten, so the Setup edits above are kept.
    n_cached = 0
    for (sheet_name, row_idx, col_idx), value in cached_values.items():
        ws = wb[sheet_name]
        cell = ws.cell(row=row_idx, column=col_idx)
        if str(cell.value).startswith('=') if cell.value else False:
            cell.value = value
            n_cached += 1

    # Save the workbook
    wb.save(interface_path)
    wb.close()

    print(f"✓ Configuration updated: Run_Base_Future=No, Run_RDM=Yes "
          f"({n_cached} formula cells cached across all sheets)")

def generate_metrics(platform_dir, metrics_file):
    """
    Generate metrics JSON file for DVC tracking.
    Extracts RDM experiment statistics.
    """
    metrics = {
        "stage": "rdm_experiment",
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "futures_generated": 0,
        "total_parquet_files": 0,
        "scenarios_processed": 0
    }

    # Count futures and files
    if platform_dir.exists():
        futures_dir = platform_dir / 'Futures'
        if futures_dir.exists():
            # Count future directories
            future_dirs = [d for d in futures_dir.iterdir() if d.is_dir() and d.name.startswith('Future_')]
            metrics["futures_generated"] = len(future_dirs)

            # Count parquet files per future
            for future_dir in future_dirs:
                parquet_files = list(future_dir.glob('*.parquet'))
                metrics["total_parquet_files"] += len(parquet_files)

                # Count unique scenarios within this future
                scenarios = set()
                for pf in parquet_files:
                    # Extract scenario name from filename pattern
                    parts = pf.stem.split('_')
                    if len(parts) >= 2:
                        scenarios.add(parts[0])
                metrics["scenarios_processed"] = len(scenarios)

    # Write metrics
    metrics_file.parent.mkdir(parents=True, exist_ok=True)
    with open(metrics_file, 'w') as f:
        json.dump(metrics, f, indent=2)

    print(f"✓ Metrics written to {metrics_file}")
    print(f"  - Futures generated: {metrics['futures_generated']}")
    print(f"  - Total parquet files: {metrics['total_parquet_files']}")
    print(f"  - Scenarios processed: {metrics['scenarios_processed']}")

def main():
    """Main execution function"""
    print("=" * 70)
    print("AFR_RDM - RDM Experiment Execution")
    print("=" * 70)

    # Paths
    project_root = Path(__file__).parent.parent
    interface_path = project_root / 'src' / 'Interface_RDM.xlsx'
    platform_dir = project_root / 'src' / 'workflow' / '1_Experiment' / 'Experimental_Platform'
    metrics_file = project_root / 'src' / 'workflow' / '1_Experiment' / 'rdm_experiment_metrics.json'

    # Verify Interface_RDM.xlsx exists
    if not interface_path.exists():
        print(f"❌ Error: {interface_path} not found")
        sys.exit(1)

    # Backup and configure
    backup_path = None
    start_time = time.time()

    try:
        # Backup original configuration
        backup_path = backup_interface(interface_path)
        print(f"✓ Backup created: {backup_path.name}")

        # Configure for RDM experiment only
        configure_for_rdm_only(interface_path)

        # Change to src directory and execute RUN_RDM
        print("\n🚀 Executing RUN_RDM.py for RDM experiment...")
        print("-" * 70)

        os.chdir(project_root / 'src')

        # Import and execute (this preserves the original RUN_RDM.py logic)
        import RUN_RDM

        print("-" * 70)

        # Generate metrics
        elapsed_time = time.time() - start_time
        print(f"\n⏱️  Execution time: {elapsed_time:.2f} seconds")

        generate_metrics(platform_dir, metrics_file)

        print("\n✅ RDM experiment completed successfully!")

    except Exception as e:
        print(f"\n❌ Error during execution: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    finally:
        # Always restore original configuration
        if backup_path:
            os.chdir(project_root)
            restore_interface(interface_path, backup_path)
            print(f"✓ Configuration restored from backup")

if __name__ == "__main__":
    main()
